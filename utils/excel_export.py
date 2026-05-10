import io
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
COMPANY_FILL  = PatternFill("solid", fgColor="E2EFDA")
COMPANY_FONT  = Font(bold=True, color="375623", size=10)
IMPORT_FILL   = PatternFill("solid", fgColor="FFF2CC")
IMPORT_FONT   = Font(bold=True, color="7F6000", size=10)
THIN          = Side(style="thin", color="BFBFBF")
THIN_BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT_FILL      = PatternFill("solid", fgColor="F2F2F2")

# Sheet 1 columns: (field_key, header_label, col_width)
MAPPING_COLUMNS = [
    ("recipient_country",  "Recipient Country",    10),
    ("transaction_type",   "Transaction Type",     14),
    ("tax_type",           "Tax Type",             12),
    ("tax_rate",           "Tax Rate",             10),
    ("supplier_location",  "Supplier Location",    14),
    ("item_nature",        "Item Nature",          12),
    ("vat_treatment",      "VAT Treatment",        30),
    ("scenario_name",      "Scenario Name",        40),
    ("default_code",       "Default Tax Code",     16),
    ("company_code",       "Company Tax Code",     16),
    ("notes",              "Notes",                30),
]

# Sheet 3 import-format columns
IMPORT_COLUMNS = [
    ("code",             "code",             16),
    ("description",      "description",      40),
    ("itemsTaxRate",     "itemsTaxRate",     14),
    ("itemsTaxRateCoupa","itemsTaxRateCoupa",18),
    ("recipientCountry", "recipientCountry", 16),
    ("vendorCountry",    "vendorCountry",    14),
    ("taxTreatment",     "taxTreatment",     30),
]


def _header_cell(ws, row, col, value, fill=HEADER_FILL, font=HEADER_FONT):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    return cell


def _fmt_rate(tax_rate) -> str:
    if tax_rate is None:
        return "—"
    return f"{tax_rate * 100:.2f}%"


def _rate_decimal(tax_rate):
    """0.19 → '0.19', 0 → '0', None → ''"""
    if tax_rate is None:
        return ""
    val = tax_rate
    return str(int(val)) if val == int(val) else str(val)


def _rate_coupa(tax_rate):
    """0.19 → 19, 0.07 → 7, 0.055 → 5.5, None → ''"""
    if tax_rate is None:
        return ""
    val = round(tax_rate * 100, 10)
    return int(val) if val == int(val) else val


def _expand_rows(rows: list[dict], eu_codes: list[str], non_eu_codes: list[str]) -> list[dict]:
    """
    Expand each scenario row into per-vendor-country rows for Sheet 3.
    EU/NonEU supplier locations are fanned out to individual country codes,
    excluding the recipient country itself.
    """
    out = []
    for r in rows:
        rc   = r["recipient_country"]
        loc  = r.get("supplier_location") or "N/A"
        code = r["company_code"] or r["default_code"]

        base = {
            "code":              code,
            "description":       r["scenario_name"],
            "itemsTaxRate":      _rate_decimal(r.get("tax_rate")),
            "itemsTaxRateCoupa": _rate_coupa(r.get("tax_rate")),
            "recipientCountry":  rc,
            "taxTreatment":      r["vat_treatment"],
        }

        if loc == "Domestic":
            out.append({**base, "vendorCountry": rc})
        elif loc == "EU":
            for cc in eu_codes:
                if cc != rc:
                    out.append({**base, "vendorCountry": cc})
        elif loc == "NonEU":
            for cc in non_eu_codes:
                if cc != rc:
                    out.append({**base, "vendorCountry": cc})
        else:
            # N/A (payroll, compensation, etc.)
            out.append({**base, "vendorCountry": "N/A"})

    return out


def build_excel(
    rows: list[dict],
    company: dict,
    eu_codes: list[str],
    non_eu_codes: list[str],
    vat_labels: dict[str, str],
    country_filter: str | None = None,
) -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Tax Code Mapping ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tax Code Mapping"
    ws1.freeze_panes = "A3"

    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(MAPPING_COLUMNS))
    tc = ws1.cell(row=1, column=1, value=f"Tax Code Mapping — {company['acronym']} — {company['name']}")
    tc.font   = Font(bold=True, color="FFFFFF", size=12)
    tc.fill   = PatternFill("solid", fgColor="1F3864")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 22

    for ci, (_, label, width) in enumerate(MAPPING_COLUMNS, start=1):
        _header_cell(ws1, 2, ci, label)
        ws1.column_dimensions[get_column_letter(ci)].width = width
    ws1.row_dimensions[2].height = 30

    for ri, record in enumerate(rows, start=3):
        is_alt = (ri % 2 == 0)
        for ci, (field, _, _) in enumerate(MAPPING_COLUMNS, start=1):
            value = record.get(field, "")
            if field == "tax_rate":
                value = _fmt_rate(value if value != "" else None)
            elif field == "vat_treatment":
                value = vat_labels.get(value, value)
            elif field in ("notes", "supplier_location", "item_nature") and not value:
                value = "—"
            elif field == "company_code" and not value:
                value = ""

            cell = ws1.cell(row=ri, column=ci, value=value)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font      = Font(size=9)

            if field == "company_code":
                cell.fill      = COMPANY_FILL
                cell.font      = COMPANY_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif field == "default_code":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font      = Font(size=9, italic=True, color="595959")
            elif is_alt:
                cell.fill = ALT_FILL

        ws1.row_dimensions[ri].height = 18

    # ── Sheet 2: Metadata ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Metadata")
    meta = [
        ("Company Acronym",  company["acronym"]),
        ("Company Name",     company["name"]),
        ("Export Date",      date.today().isoformat()),
        ("Country Filter",   country_filter or "All"),
        ("Total Scenarios",  len(rows)),
        ("Assigned Codes",   sum(1 for r in rows if r.get("company_code"))),
    ]
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 30
    for ri, (key, val) in enumerate(meta, start=1):
        ws2.cell(row=ri, column=1, value=key).font = Font(bold=True)
        ws2.cell(row=ri, column=2, value=val).alignment = Alignment(horizontal="left")

    # ── Sheet 3: Import Format (expanded per vendor country) ───────────────
    ws3 = wb.create_sheet("Import Format")
    ws3.freeze_panes = "A2"

    for ci, (_, label, width) in enumerate(IMPORT_COLUMNS, start=1):
        _header_cell(ws3, 1, ci, label, fill=IMPORT_FILL, font=IMPORT_FONT)
        ws3.column_dimensions[get_column_letter(ci)].width = width
    ws3.row_dimensions[1].height = 26

    expanded = _expand_rows(rows, eu_codes, non_eu_codes)
    for ri, rec in enumerate(expanded, start=2):
        is_alt = (ri % 2 == 0)
        for ci, (field, _, _) in enumerate(IMPORT_COLUMNS, start=1):
            val  = rec.get(field, "")
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            cell.font      = Font(size=9)
            if field == "code":
                cell.fill      = COMPANY_FILL
                cell.font      = COMPANY_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif is_alt:
                cell.fill = ALT_FILL
        ws3.row_dimensions[ri].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
